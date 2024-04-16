from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import sqlite3 as sq
from selenium.webdriver.chrome.service import Service
import openpyxl


def export():
    clear()
    wb = Workbook()
    ws = wb.active
    ws.title = "Winter_shoes"
    ws.append(['Title', 'Price', 'Image', 'Link', ''])

    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    s = Service('C:/Users/denis/Downloads/chromedriver.exe')
    driver = webdriver.Chrome(service=s)
    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win64",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )
    url = "https://sneakerhead.ru/shoes/winter-shoes/"
    driver.get(url)
    blocks = driver.find_element(By.CLASS_NAME, "product-cards__list")
    posts = blocks.find_elements(By.CLASS_NAME, "product-cards__item")
    for post in posts:
        title = post.find_element(By.CLASS_NAME, "product-card").find_element(By.CLASS_NAME,
                                                                              "product-card__title").find_element(
            By.CLASS_NAME, "product-card__link").get_attribute("title")
        price = post.find_element(By.CLASS_NAME, "product-card").find_element(By.CLASS_NAME,
                                                                              "product-card__price").find_element(
            By.CLASS_NAME, "product-card__price-value").text.replace(" ", ".")
        image = "https://sneakerhead.ru/" + post.find_element(By.CLASS_NAME, "product-card").find_element(By.CLASS_NAME,
                                                                                                          "product-card__image").find_element(
            By.CLASS_NAME, "product-card__image-inner").find_element(By.TAG_NAME, "source").get_attribute(
            "data-srcset").split("1x")[0]
        link = post.find_element(By.CLASS_NAME, "product-card").find_element(By.CLASS_NAME,
                                                                             "product-card__title").find_element(
            By.CLASS_NAME, "product-card__link").get_attribute("href")
        ws.append([title, price, image, link])
        print(title, price, image, link)

    wb.save('wintershoes.xlsx')
    wb.close()

    con = sq.connect("sneakers.db")
    cur = con.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS winter_shoes (
    title TEXT,
    price TEXT,
    image TEXT,
    link TEXT
    )""")

    book = openpyxl.open("wintershoes.xlsx", read_only=True)
    sheet = book.active
    for row in range(2, sheet.max_row + 1):
        data = []
        for col in range(1, 5):
            value = sheet.cell(row, col).value
            data.append(value)
        cur.execute("INSERT INTO winter_shoes VALUES (?, ?, ?, ?);", (data[0], data[1], data[2], data[3]))

    con.commit()
    con.close()


def clear():
    con = sq.connect("sneakers.db")
    cur = con.cursor()
    cur.execute("DELETE FROM winter_shoes")
    con.commit()
    con.close()
